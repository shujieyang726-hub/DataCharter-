#!/usr/bin/env python3
# scripts/chart_recommender.py
"""
PPT Master - 数据匹配图表推荐工具

根据用户提交的数据（JSON/CSV/Excel）自动推荐最合适的图表类型。

Usage:
    python3 scripts/chart_recommender.py data.json
    python3 scripts/chart_recommender.py sales.csv
    python3 scripts/chart_recommender.py report.xlsx -o recommendation.json
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

# 时间模式正则
TIME_PATTERNS = [
    re.compile(r"^\d{1,2}月$"),           # 1月, 12月
    re.compile(r"^Q[1-4]$", re.I),        # Q1-Q4
    re.compile(r"^\d{4}$"),               # 2024
    re.compile(r"^\d{4}[-/]\d{1,2}$"),    # 2024-01, 2024/1
    re.compile(r"^\d{4}年$"),             # 2024年
    re.compile(r"^第[一二三四]季度$"),      # 第一季度
    re.compile(r"^(周|星期)[一二三四五六日]$"),  # 周一
    re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)", re.I),
]


def _is_time_label(label: str) -> bool:
    """判断标签是否为时间标签"""
    return any(p.match(label.strip()) for p in TIME_PATTERNS)


def _is_time_series_labels(labels: List[str]) -> bool:
    """判断标签列表是否为时间序列"""
    if len(labels) < 2:
        return False
    time_count = sum(1 for l in labels if _is_time_label(l))
    return time_count / len(labels) >= 0.5


def analyze_dataset(dataset: Dict[str, Any]) -> Dict[str, Any]:
    """分析单个数据集，提取特征"""
    categories = dataset.get("categories") or dataset.get("labels") or []
    values = dataset.get("values") or []
    x_axis = dataset.get("x_axis") or []
    series_list = dataset.get("series") or []
    unit = dataset.get("unit", "")

    if series_list:
        series_count = len(series_list)
        all_values = []
        for s in series_list:
            all_values.extend(s.get("values", []))
        if not categories and x_axis:
            categories = x_axis
    else:
        series_count = 1
        all_values = list(values)

    category_count = len(categories) or len(x_axis)

    labels_to_check = x_axis if x_axis else categories
    is_time_series = _is_time_series_labels(labels_to_check)

    is_percentage = False
    if "%" in unit:
        is_percentage = True
    elif values and all(isinstance(v, (int, float)) for v in values):
        total = sum(abs(v) for v in values)
        if 95 <= total <= 105 and len(values) >= 2:
            is_percentage = True

    has_negative = any(v < 0 for v in all_values if isinstance(v, (int, float)))

    numeric_values = [v for v in all_values if isinstance(v, (int, float))]
    value_range = (min(numeric_values), max(numeric_values)) if numeric_values else (0, 0)

    label_length = (sum(len(str(c)) for c in categories) / len(categories)) if categories else 0

    if is_percentage:
        data_intent = "composition"
    elif is_time_series:
        data_intent = "trend"
    elif has_negative:
        data_intent = "comparison"
    elif series_count == 1 and category_count >= 2:
        data_intent = "comparison"
    elif series_count >= 2:
        data_intent = "comparison"
    else:
        data_intent = "comparison"

    return {
        "data_intent": data_intent,
        "category_count": category_count,
        "series_count": series_count,
        "is_time_series": is_time_series,
        "is_percentage": is_percentage,
        "has_negative": has_negative,
        "value_range": value_range,
        "label_length": round(label_length, 1),
    }


# ─── 推荐决策树 ───


def recommend_chart(features: Dict[str, Any]) -> Dict[str, Any]:
    """根据数据特征推荐图表类型"""
    is_pct = features["is_percentage"]
    is_ts = features["is_time_series"]
    cat_count = features["category_count"]
    series_count = features["series_count"]
    has_neg = features["has_negative"]
    label_len = features["label_length"]

    if is_pct and cat_count <= 6:
        return {
            "primary": {"type": "donut_chart", "confidence": 0.95,
                        "reason": f"占比数据，{cat_count}个分类，适合环形图"},
            "alternatives": [
                {"type": "pie_chart", "confidence": 0.80,
                 "reason": "饼图也适合占比展示"},
            ],
        }

    if is_ts:
        if series_count == 1:
            return {
                "primary": {"type": "line_chart", "confidence": 0.92,
                            "reason": "单系列时间序列，折线图最直观"},
                "alternatives": [
                    {"type": "area_chart", "confidence": 0.75,
                     "reason": "面积图可强调趋势下方区域"},
                ],
            }
        if series_count == 2:
            return {
                "primary": {"type": "dual_axis_line_chart", "confidence": 0.88,
                            "reason": "双系列时间序列，双轴折线图对比清晰"},
                "alternatives": [
                    {"type": "line_chart", "confidence": 0.70,
                     "reason": "单轴折线图也可展示双系列"},
                ],
            }
        return {
            "primary": {"type": "stacked_area_chart", "confidence": 0.82,
                        "reason": f"{series_count}个系列的时间趋势，堆叠面积图展示总量与构成"},
            "alternatives": [
                {"type": "line_chart", "confidence": 0.65,
                 "reason": "多折线对比"},
            ],
        }

    if has_neg:
        return {
            "primary": {"type": "waterfall_chart", "confidence": 0.85,
                        "reason": "包含负数值，瀑布图展示增减变化"},
            "alternatives": [
                {"type": "butterfly_chart", "confidence": 0.65,
                 "reason": "蝴蝶图可做双向对比"},
            ],
        }

    if series_count == 1 and cat_count <= 8:
        if label_len > 6:
            return {
                "primary": {"type": "horizontal_bar_chart", "confidence": 0.90,
                            "reason": f"标签较长（平均{label_len:.0f}字符），横向柱状图更清晰"},
                "alternatives": [
                    {"type": "bar_chart", "confidence": 0.65,
                     "reason": "纵向柱状图（标签可能截断）"},
                ],
            }
        return {
            "primary": {"type": "bar_chart", "confidence": 0.92,
                        "reason": f"单系列{cat_count}个类别对比"},
            "alternatives": [
                {"type": "horizontal_bar_chart", "confidence": 0.70,
                 "reason": "横向展示也可"},
                {"type": "radar_chart", "confidence": 0.40,
                 "reason": "雷达图可做多维对比"},
            ],
        }

    if series_count >= 2 and not is_ts:
        return {
            "primary": {"type": "grouped_bar_chart", "confidence": 0.88,
                        "reason": f"{series_count}个系列的分组对比"},
            "alternatives": [
                {"type": "stacked_bar_chart", "confidence": 0.72,
                 "reason": "堆叠柱状图展示构成"},
            ],
        }

    if cat_count <= 4 and series_count == 1:
        return {
            "primary": {"type": "kpi_cards", "confidence": 0.75,
                        "reason": f"仅{cat_count}个指标，KPI卡片更直观"},
            "alternatives": [
                {"type": "bar_chart", "confidence": 0.60,
                 "reason": "柱状图也可展示少量类别"},
            ],
        }

    return {
        "primary": {"type": "bar_chart", "confidence": 0.50,
                    "reason": "通用兜底推荐"},
        "alternatives": [],
    }


# ─── 布局建议 ───

LAYOUT_MAP = {
    1: "single",
    2: "two_column",
    3: "three_column",
    4: "grid_2x2",
    5: "grid_top2_bottom3",
    6: "grid_2x3",
}


def suggest_layout(recommendations: List[Dict[str, Any]]) -> Dict[str, Any]:
    """根据推荐结果建议布局"""
    n = len(recommendations)
    layout_type = LAYOUT_MAP.get(n, "grid_2x3" if n > 6 else "single")
    slots = [
        {"index": i, "chart_type": r["primary"]["type"]}
        for i, r in enumerate(recommendations)
    ]
    return {
        "total_charts": n,
        "layout_type": layout_type,
        "slots": slots,
    }


# ─── 数据加载 ───


def load_json_data(file_path: Path) -> List[Dict[str, Any]]:
    """加载JSON数据文件"""
    data = json.loads(file_path.read_text(encoding="utf-8"))
    if "datasets" in data:
        return data["datasets"]
    if isinstance(data, list):
        return data
    return [data]


def load_csv_data(file_path: Path) -> List[Dict[str, Any]]:
    """加载CSV文件，自动推断列角色"""
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    if not rows:
        return []

    headers = list(rows[0].keys())
    numeric_cols = []
    category_col = None

    for h in headers:
        vals = [r[h] for r in rows]
        try:
            [float(v) for v in vals if v.strip()]
            numeric_cols.append(h)
        except ValueError:
            if category_col is None:
                category_col = h

    if not numeric_cols:
        return []

    categories = [r[category_col] for r in rows] if category_col else [str(i) for i in range(len(rows))]

    if len(numeric_cols) == 1:
        values = [float(r[numeric_cols[0]]) for r in rows if r[numeric_cols[0]].strip()]
        return [{
            "title": file_path.stem,
            "categories": categories,
            "values": values,
        }]

    series = []
    for col in numeric_cols:
        series.append({
            "name": col,
            "values": [float(r[col]) for r in rows if r[col].strip()],
        })
    return [{
        "title": file_path.stem,
        "x_axis": categories,
        "series": series,
    }]


def load_excel_data(file_path: Path) -> List[Dict[str, Any]]:
    """加载Excel文件"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[ERROR] 需要 openpyxl 库支持Excel: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    datasets = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:]

        numeric_cols = []
        category_col = None

        for ci, h in enumerate(headers):
            vals = [r[ci] for r in data_rows if r[ci] is not None]
            if vals and all(isinstance(v, (int, float)) for v in vals):
                numeric_cols.append((ci, h))
            elif category_col is None:
                category_col = (ci, h)

        if not numeric_cols:
            continue

        cats = ([str(r[category_col[0]]) for r in data_rows if r[category_col[0]] is not None]
                if category_col else [str(i) for i in range(len(data_rows))])

        if len(numeric_cols) == 1:
            ci, name = numeric_cols[0]
            datasets.append({
                "title": sheet,
                "categories": cats,
                "values": [float(r[ci]) for r in data_rows if r[ci] is not None],
            })
        else:
            series = []
            for ci, name in numeric_cols:
                series.append({
                    "name": name,
                    "values": [float(r[ci]) for r in data_rows if r[ci] is not None],
                })
            datasets.append({
                "title": sheet,
                "x_axis": cats,
                "series": series,
            })

    wb.close()
    return datasets


# ─── CLI ───


def main() -> int:
    parser = argparse.ArgumentParser(
        description="PPT Master - 数据匹配图表推荐工具",
    )
    parser.add_argument("data_file", type=Path, help="数据文件（JSON/CSV/Excel）")
    parser.add_argument("-o", "--output", type=Path, default=None,
                        help="输出推荐结果JSON（默认: 打印到终端）")
    args = parser.parse_args()

    data_path = args.data_file.expanduser().resolve()
    if not data_path.exists():
        print(f"[ERROR] 文件不存在: {data_path}")
        return 1

    suffix = data_path.suffix.lower()
    if suffix == ".json":
        datasets = load_json_data(data_path)
    elif suffix == ".csv":
        datasets = load_csv_data(data_path)
    elif suffix in (".xlsx", ".xls"):
        datasets = load_excel_data(data_path)
    else:
        print(f"[ERROR] 不支持的文件类型: {suffix}")
        return 1

    if not datasets:
        print("[ERROR] 未找到有效数据集")
        return 1

    recommendations = []
    for i, ds in enumerate(datasets):
        features = analyze_dataset(ds)
        rec = recommend_chart(features)
        recommendations.append({
            "dataset_index": i,
            "dataset_title": ds.get("title", f"dataset_{i}"),
            "features": features,
            **rec,
        })

    layout = suggest_layout(recommendations)

    output = {
        "recommendations": recommendations,
        "suggested_layout": layout,
    }

    if args.output:
        out_path = args.output.expanduser().resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(
            json.dumps(output, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"[OK] 推荐结果已保存: {out_path}")
    else:
        print(json.dumps(output, ensure_ascii=False, indent=2))

    print(f"\n[OK] 分析了 {len(datasets)} 个数据集，建议布局: {layout['layout_type']}")
    for r in recommendations:
        print(f"  #{r['dataset_index']}: {r['dataset_title']} → {r['primary']['type']} ({r['primary']['confidence']:.0%})")

    return 0


if __name__ == "__main__":
    sys.exit(main())
